from urllib.parse import quote
from typing import List, Optional
import uuid
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Response, Query
from datetime import datetime, date, timedelta
import pytz
import openpyxl
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
from pydantic import BaseModel, Field

from app.database import get_db
from app.models import TreasuryNotification, Request, User, Import, ApprovalProcess
from app.schemas import RequestResponse, ImportType, Category
from app.auth import get_current_user, require_treasury
from app.utils.excel_processor import process_excel_file

from typing import Optional, List

# Модель для узла дерева
class TreeNode(BaseModel):
    id: str
    name: str
    type: str  # 'root', 'organization', 'department', 'user', 'import'
    count: int
    amount: float
    children: Optional[List['TreeNode']] = None
    import_id: Optional[str] = None
    user_id: Optional[str] = None
    organization: Optional[str] = None
    department: Optional[str] = None

TreeNode.update_forward_refs()

# Модель для комментариев процессов согласования
class ApprovalCommentResponse(BaseModel):
    """Ответ с комментариями процесса согласования"""
    has_comment: bool = Field(..., description="Есть ли комментарий")
    treasury_comment: Optional[str] = Field(None, description="Комментарий казначейства")
    approval_process_id: Optional[UUID] = Field(None, description="ID процесса согласования")
    comment: Optional[str] = Field(None, description="Комментарий заместителя (для казначейства)")

router = APIRouter()

# Модель для запроса экспорта
class ExportRequest(BaseModel):
    request_ids: Optional[List[uuid.UUID]] = None
    export_all: bool = False

# Создаем схему для ответа с уведомлениями
from typing import Optional as Opt
from uuid import UUID

class TreasuryNotificationResponse(BaseModel):
    id: UUID
    approval_process_id: Opt[UUID]
    category: str
    deputy_name: str
    comment: str
    request_count: int
    total_amount: float
    is_read: bool
    created_at: datetime

    class Config:
        from_attributes = True

@router.get("/notifications", response_model=List[TreasuryNotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение уведомлений от заместителя ГД
    """
    query = db.query(TreasuryNotification)

    if unread_only:
        query = query.filter(TreasuryNotification.is_read == False)

    notifications = query.order_by(TreasuryNotification.created_at.desc()).all()

    return notifications

@router.post("/notifications/{notification_id}/mark-read")
async def mark_notification_read(
    notification_id: uuid.UUID,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Отметка уведомления как прочитанного
    """
    notification = db.query(TreasuryNotification).filter(
        TreasuryNotification.id == notification_id
    ).first()

    if not notification:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Уведомление не найдено"
        )

    notification.is_read = True
    db.commit()

    return {"message": "Уведомление отмечено как прочитанное"}

@router.get("/for-payment", response_model=List[RequestResponse])
async def get_requests_for_payment(
    category: Category = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение заявок к оплате
    """
    query = db.query(Request).filter(Request.status == "for_payment")

    if category:
        query = query.filter(Request.category == category)

    if start_date:
        query = query.filter(Request.created_at >= start_date)

    if end_date:
        query = query.filter(Request.created_at <= end_date)

    requests = query.order_by(
        Request.priority.desc() if Request.priority else None,
        Request.created_at
    ).all()

    return requests


@router.get("/approved-for-payment", response_model=List[RequestResponse])
async def get_approved_requests_for_payment(
    category: Category = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение заявок, согласованных в оплату
    """
    query = db.query(Request).filter(Request.status == "approved_for_payment")

    if category:
        query = query.filter(Request.category == category)

    if start_date:
        query = query.filter(Request.created_at >= start_date)

    if end_date:
        query = query.filter(Request.created_at <= end_date)

    requests = query.order_by(
        Request.priority.desc() if Request.priority else None,
        Request.created_at
    ).all()

    return requests

@router.post("/export")
async def export_requests(
    export_data: ExportRequest,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Экспорт заявок в Excel
    """
    import logging
    logger = logging.getLogger(__name__)

    logger.info(f"Export request: export_all={export_data.export_all}, request_ids={export_data.request_ids}")

    query = db.query(Request).filter(Request.status == "for_payment")

    logger.info(f"Initial query count: {query.count()}")

    if not export_data.export_all and export_data.request_ids:
        logger.info(f"Filtering by request_ids: {export_data.request_ids}")
        query = query.filter(Request.id.in_(export_data.request_ids))

    requests = query.all()
    logger.info(f"Found {len(requests)} requests for export")

    if not requests:
        logger.warning("No requests found for export")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Нет заявок для экспорта"
        )

    # Создание Excel файла
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Заявки к оплате"

    # Заголовки
    headers = [
        "Статья ДДС", "Сумма", "Получатель", "Номер заявки",
        "Дата заявки", "Статус", "Организация", "Подразделение",
        "Приоритет", "Назначение", "Дата оплаты", "Заявитель"
    ]

    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Данные
    for row_num, request in enumerate(requests, 2):
        worksheet.cell(row=row_num, column=1, value=request.article)
        # Форматируем сумму как число с разделителями
        cell = worksheet.cell(row=row_num, column=2, value=request.amount)
        cell.number_format = '#,##0.00'
        worksheet.cell(row=row_num, column=3, value=request.recipient)
        worksheet.cell(row=row_num, column=4, value=request.request_number)
        # Конвертируем время в Московский часовой пояс
        moscow_tz = pytz.timezone('Europe/Moscow')
        if request.request_date.tzinfo is None:
            # Если время без часового пояса, считаем что это UTC
            request_date_utc = pytz.utc.localize(request.request_date)
        else:
            request_date_utc = request.request_date.astimezone(pytz.utc)
        request_date_moscow = request_date_utc.astimezone(moscow_tz)
        worksheet.cell(row=row_num, column=5, value=request_date_moscow.strftime("%d.%m.%Y %H:%M:%S"))
        worksheet.cell(row=row_num, column=6, value=request.status)
        worksheet.cell(row=row_num, column=7, value=request.organization)
        worksheet.cell(row=row_num, column=8, value=request.department)
        worksheet.cell(row=row_num, column=9, value=request.priority)
        worksheet.cell(row=row_num, column=10, value=request.purpose)
        worksheet.cell(row=row_num, column=11, value=request.payment_date.strftime("%d.%m.%Y") if request.payment_date else "")
        worksheet.cell(row=row_num, column=12, value=request.applicant)

    # Сохранение в буфер
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    # Создание имени файла
    filename = f"Заявки_к_оплате_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"

    logger.info(f"Export completed: {len(requests)} requests, filename: {filename}")

    # Возвращаем файл как ответ
    return Response(
        content=excel_buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{filename}\"",
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='')}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8"
        }
    )

@router.post("/special-import")
async def special_import(
    import_type: ImportType = Form(...),
    payment_date: str = Form(...),
    file: UploadFile = File(...),
    comment: str = Form(None),
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Импорт особых заявок (непереносимые оплаты, графики, согласовано в оплату)
    """
    # Валидация файла
    if file.size > 5 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл превышает лимит 5 МБ"
        )

    # Создание записи об импорте
    db_import = Import(
        user_id=current_user.id,
        file_name=file.filename,
        file_size=file.size,
        payment_date=datetime.strptime(payment_date, "%Y-%m-%d").date(),
        import_type=import_type.value,
        comment=comment,
        status="processing"
    )

    db.add(db_import)
    db.commit()
    db.refresh(db_import)

    imported_count = 0

    try:
        # Чтение файла
        content = await file.read()

        # Обработка Excel файла
        requests_data = process_excel_file(content)

        for request_data in requests_data:
            # Создание особой заявки
            # Маппинг ImportType -> Category и treasury_import_type
            category_map = {
                "non_transferable": "non_transferable",
                "schedules": "schedules",
                "approved_by_director": "schedules",
                "approved_for_payment": "schedules",
            }
            treasury_import_type_map = {
                "non_transferable": "non_transferable",
                "schedules": "graphs",
                "approved_by_director": "approved_by_director",
                "approved_for_payment": "approved_by_director",
            }
            
            db_request = Request(
                article=request_data.get("Статья ДДС", ""),
                amount=float(request_data.get("Сумма", 0)),
                recipient=request_data.get("Получатель", ""),
                request_number=request_data.get("Номер заявки", ""),
                request_date=pytz.timezone('Europe/Moscow').localize(datetime.strptime(
                    request_data.get("Дата заявки", datetime.now().strftime("%d.%m.%Y %H:%M:%S")),
                    "%d.%m.%Y %H:%M:%S"
                )),
                organization=request_data.get("Организация", ""),
                department=request_data.get("Подразделение", ""),
                purpose=request_data.get("Назначение", ""),
                applicant=request_data.get("Заявитель", ""),
                category=category_map.get(import_type.value, "schedules"),
                import_type="special",
                treasury_import_type=treasury_import_type_map.get(import_type.value, import_type.value),
                created_by=current_user.id,
                import_id=db_import.id,
                source='treasury',
                status="pending"
            )

            db.add(db_request)
            # Сохраняем каждую заявку
            db.flush()
            
            # Автоматическая категоризация
            from app.utils.categorization import categorize_request
            categorize_request(db_request, db)
            
            imported_count += 1

        # Обновление статуса импорта
        db_import.status = "completed"
        db_import.imported_count = imported_count
        db.commit()

    except Exception as e:
        # В случае ошибки
        db_import.status = "failed"
        db_import.error_message = str(e)
        db.commit()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Ошибка при импорте: {str(e)}"
        )

    return {
        "message": f"Импорт завершен успешно",
        "imported_count": imported_count,
        "import_type": import_type.value
    }
async def update_payment_date(
    request_ids: List[uuid.UUID],
    payment_date: date,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Обновление даты оплаты для выбранных заявок
    """
    requests = db.query(Request).filter(
        Request.id.in_(request_ids),
        Request.status.in_(["for_payment", "approved_for_payment"])
    ).all()

    if not requests:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Заявки не найдены или не могут быть оплачены"
        )

    # Обновление даты оплаты
    for request in requests:
        request.payment_date = payment_date
        request.updated_at = datetime.utcnow()

    db.commit()

    return {
        "message": f"Дата оплаты обновлена для {len(requests)} заявок",
        "count": len(requests),
        "payment_date": payment_date.strftime("%d.%m.%Y")
    }

@router.get("/statistics")
async def get_treasury_statistics(
    period: str = "month",  # day, week, month, year
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение статистики для казначейства
    """
    # Расчет дат в зависимости от периода
    end_date = datetime.now()

    if period == "day":
        start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start_date = end_date - timedelta(days=7)
    elif period == "month":
        start_date = end_date - timedelta(days=30)
    elif period == "year":
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)

    # Статистика по статусам
    status_stats = {}
    for status in ["for_payment", "approved_for_payment", "rejected"]:
        query = db.query(Request).filter(
            Request.status == status,
            Request.updated_at >= start_date,
            Request.updated_at <= end_date
        )

        count = query.count()

        if count > 0:
            total_amount = db.query(func.sum(Request.amount)).filter(
                Request.status == status,
                Request.updated_at >= start_date,
                Request.updated_at <= end_date
            ).scalar() or 0

            status_stats[status] = {
                "count": count,
                "total_amount": total_amount
            }

    # Статистика по категориям
    category_stats = {}
    categories = db.query(Request.category).distinct().all()

    for category_tuple in categories:
        category = category_tuple[0]
        query = db.query(Request).filter(
            Request.category == category,
            Request.status.in_(["for_payment", "approved_for_payment"]),
            Request.updated_at >= start_date,
            Request.updated_at <= end_date
        )

        count = query.count()

        if count > 0:
            total_amount = db.query(func.sum(Request.amount)).filter(
                Request.category == category,
                Request.status.in_(["for_payment", "approved_for_payment"]),
                Request.updated_at >= start_date,
                Request.updated_at <= end_date
            ).scalar() or 0

            category_stats[category] = {
                "count": count,
                "total_amount": total_amount
            }

    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "status_statistics": status_stats,
        "category_statistics": category_stats,
        "total_requests": sum([stats["count"] for stats in status_stats.values()]),
        "total_amount": sum([stats["total_amount"] for stats in status_stats.values()])
    }

# Модель для отправки заявок на согласование
class SendToApprovalRequest(BaseModel):
    request_ids: List[uuid.UUID]

@router.post("/send-to-approval")
async def send_requests_to_approval(
    request_data: SendToApprovalRequest,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Отправка заявок на согласование заместителю ГД
    """
    # Получаем заявки
    requests = db.query(Request).filter(
        Request.id.in_(request_data.request_ids),
        Request.status.in_(["draft", "pending"])
    ).all()

    if not requests:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Не найдено заявок со статусом 'Черновик'"
        )

    # Проверяем, что все заявки принадлежат казначейству (source = 'treasury')
    for request in requests:
        if request.source != 'treasury':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Заявка {request.id} не из казначейства"
            )

    # Обновляем статус на 'pending' (на согласовании)
    for request in requests:
        request.status = "approved_for_payment"
        request.updated_at = datetime.utcnow()

    # Собираем информацию для уведомления
    categories = list(set([req.category for req in requests]))
    total_amount = sum([req.amount for req in requests])
    request_count = len(requests)

    # Находим заместителей (всех активных заместителей)
    deputies = db.query(User).filter(
        User.role == "deputy_director",
        User.is_active == True
    ).all()

    if not deputies:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Не найдено активных заместителей ГД"
        )

    # Импортируем функцию для создания уведомлений
    from app.routes.notifications import create_batch_for_approval_notification

    # Создаем уведомления для каждого заместителя
    for deputy in deputies:
        create_batch_for_approval_notification(
            db=db,
            deputy_id=deputy.id,
            import_id=None,  # Нет импорта, так как отправка вручную
            request_count=request_count,
            categories=categories,
            total_amount=total_amount,
            imported_by_user=current_user
        )

    db.commit()

    return {
        "message": f"{request_count} заявок отправлено на согласование",
        "request_count": request_count,
        "categories": categories,
        "total_amount": total_amount
    }

@router.get("/requests", response_model=List[RequestResponse])
async def get_treasury_requests(
    status: Optional[str] = None,
    category: Category = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение заявок казначейства с фильтрацией по статусу
    """
    query = db.query(Request).filter(Request.source == 'treasury')

    # Фильтрация по статусу
    if status:
        query = query.filter(Request.status == status)

    if category:
        query = query.filter(Request.category == category)

    if start_date:
        query = query.filter(Request.created_at >= start_date)

    if end_date:
        query = query.filter(Request.created_at <= end_date)

    # Сортировка по дате создания (новые сверху)
    requests = query.order_by(Request.created_at.desc()).all()

    return requests

# Новые endpoint'ы для заявок на согласовании в казначействе

@router.get("/pending/imports-tree")
async def get_pending_imports_tree(
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение дерева импортов с комментариями для рабочей области казначейства
    Группировка: Организация -> Подразделение -> Импорт (пользователь + комментарий)
    """
    # Получаем все импорты с хотя бы одной заявкой в статусе 'pending'
    imports_with_pending = db.query(Import).join(Request, Request.import_id == Import.id)\
        .filter(Request.status == 'pending')\
        .group_by(Import.id)\
        .all()
    
    # Также получаем одиночные заявки без импорта
    users_with_single_requests = db.query(User).join(Request, Request.created_by == User.id)\
        .filter(
            Request.status == 'pending',
            Request.import_id.is_(None)
        )\
        .distinct()\
        .all()
    
    # Собираем данные
    organizations_dict = {}
    
    # 1. Обрабатываем импорты
    for import_item in imports_with_pending:
        user = db.query(User).filter(User.id == import_item.user_id).first()
        if not user:
            continue
            
        org_name = user.organization
        dept_name = user.department
        
        if org_name not in organizations_dict:
            organizations_dict[org_name] = {}
        
        if dept_name not in organizations_dict[org_name]:
            organizations_dict[org_name][dept_name] = []
        
        # Количество заявок в этом импорте со статусом pending
        pending_count = db.query(Request).filter(
            Request.import_id == import_item.id,
            Request.status == 'pending'
        ).count()
        
        # Общая сумма заявок в этом импорте
        total_amount_result = db.query(func.sum(Request.amount)).filter(
            Request.import_id == import_item.id,
            Request.status == 'pending'
        ).first()
        total_amount = total_amount_result[0] or 0
        
        import_data = {
            'id': str(import_item.id),
            'type': 'import',
            'user_id': str(user.id),
            'user_name': user.full_name,
            'comment': import_item.comment,
            'file_name': import_item.file_name,
            'pending_count': pending_count,
            'total_amount': float(total_amount),
            'created_at': import_item.created_at.isoformat() if import_item.created_at else None
        }
        
        organizations_dict[org_name][dept_name].append(import_data)
    
    # 2. Обрабатываем одиночные заявки без импорта
    for user in users_with_single_requests:
        org_name = user.organization
        dept_name = user.department
        
        if org_name not in organizations_dict:
            organizations_dict[org_name] = {}
        
        if dept_name not in organizations_dict[org_name]:
            organizations_dict[org_name][dept_name] = []
        
        # Количество одиночных заявок пользователя
        single_count = db.query(Request).filter(
            Request.created_by == user.id,
            Request.status == 'pending',
            Request.import_id.is_(None)
        ).count()
        
        # Общая сумма одиночных заявок
        total_amount_result = db.query(func.sum(Request.amount)).filter(
            Request.created_by == user.id,
            Request.status == 'pending',
            Request.import_id.is_(None)
        ).first()
        total_amount = total_amount_result[0] or 0
        
        user_data = {
            'id': str(user.id),
            'type': 'user',
            'user_id': str(user.id),
            'user_name': user.full_name,
            'comment': None,
            'file_name': None,
            'pending_count': single_count,
            'total_amount': float(total_amount),
            'created_at': None
        }
        
        organizations_dict[org_name][dept_name].append(user_data)
    
    # Преобразуем в формат для фронтенда
    result = []
    for org_name, departments in organizations_dict.items():
        org_data = {
            'organization': org_name,
            'departments': []
        }
        
        for dept_name, imports_list in departments.items():
            # Сортируем по дате создания (новые сверху)
            sorted_imports = sorted(
                imports_list, 
                key=lambda x: x['created_at'] or '', 
                reverse=True
            )
            
            dept_data = {
                'department': dept_name,
                'imports': sorted_imports
            }
            org_data['departments'].append(dept_data)
        
        result.append(org_data)
    
    return result



@router.get("/pending/filter-by-node")
async def filter_requests_by_node(
    node_id: str,
    node_type: str,
    organization: Optional[str] = None,
    department: Optional[str] = None,
    user_id: Optional[str] = None,
    import_id: Optional[str] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Фильтрация заявок по выбранному узлу дерева
    """
    query = db.query(Request).filter(Request.status == 'pending')
    
    # Применяем фильтры в зависимости от типа узла
    if node_type == 'root':
        # Все заявки - без фильтров
        pass
    elif node_type == 'organization' and organization:
        query = query.filter(Request.organization == organization)
    elif node_type == 'department' and organization and department:
        query = query.filter(
            Request.organization == organization,
            Request.department == department
        )
    elif node_type == 'user' and user_id:
        query = query.filter(Request.created_by == uuid.UUID(user_id))
    elif node_type == 'import' and import_id:
        # Для import_id может быть как UUID импорта, так и user_id для одиночных заявок
        try:
            import_uuid = uuid.UUID(import_id)
            query = query.filter(Request.import_id == import_uuid)
        except ValueError:
            # Если не UUID, то это user_id для одиночных заявок
            query = query.filter(
                Request.created_by == uuid.UUID(import_id),
                Request.import_id.is_(None)
            )
    
    requests = query.order_by(Request.created_at.desc()).all()
    
    # Преобразуем в Response
    return [RequestResponse.from_orm(req) for req in requests]


@router.post("/pending/pivot-by-node")
async def get_pivot_by_node(
    data: dict,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение сводной таблицы для выбранного узла дерева
    """
    node_type = data.get('node_type')
    organization = data.get('organization')
    department = data.get('department')
    user_id = data.get('user_id')
    import_id = data.get('import_id')
    
    query = db.query(Request).filter(Request.status == 'pending')
    
    # Применяем фильтры в зависимости от типа узла
    if node_type == 'organization' and organization:
        query = query.filter(Request.organization == organization)
    elif node_type == 'department' and organization and department:
        query = query.filter(
            Request.organization == organization,
            Request.department == department
        )
    elif node_type == 'user' and user_id:
        query = query.filter(Request.created_by == uuid.UUID(user_id))
    elif node_type == 'import' and import_id:
        try:
            import_uuid = uuid.UUID(import_id)
            query = query.filter(Request.import_id == import_uuid)
        except ValueError:
            query = query.filter(
                Request.created_by == uuid.UUID(import_id),
                Request.import_id.is_(None)
            )
    # Для root не применяем фильтры - все заявки
    
    requests = query.all()
    
    # Собираем данные для сводной таблицы
    pivot_data = {}
    departments_set = set()
    
    for request in requests:
        org = request.organization
        dept = request.department
        
        departments_set.add(dept)
        
        if org not in pivot_data:
            pivot_data[org] = {}
        
        if dept not in pivot_data[org]:
            pivot_data[org][dept] = 0
        
        pivot_data[org][dept] += request.amount
    
    # Преобразуем в формат для фронтенда
    departments = sorted(list(departments_set))
    rows = []
    
    for org, depts in pivot_data.items():
        row = {
            'organization': org,
            'departments': {}
        }
        
        for dept in departments:
            row['departments'][dept] = depts.get(dept, 0)
        
        rows.append(row)
    
    # Сортируем строки по названию организации
    rows.sort(key=lambda x: x['organization'])
    
    return {
        'departments': departments,
        'rows': rows
    }
@router.get("/pending/requests")

@router.get("/pending/tree")
async def get_pending_tree(
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение дерева заявок для навигации
    Структура: Все заявки -> Организации -> Подразделения -> Пользователи/Импорты
    """
    # Получаем все заявки в статусе 'pending'
    all_requests = db.query(Request).filter(Request.status == 'pending').all()
    
    # Собираем данные для дерева
    tree_data = {}
    
    # Обрабатываем все заявки
    for request in all_requests:
        org = request.organization or "Без организации"
        dept = request.department or "Без подразделения"
        
        if org not in tree_data:
            tree_data[org] = {
                'name': org,
                'departments': {},
                'total_count': 0,
                'total_amount': 0
            }
        
        if dept not in tree_data[org]['departments']:
            tree_data[org]['departments'][dept] = {
                'name': dept,
                'users': {},
                'total_count': 0,
                'total_amount': 0
            }
        
        # Определяем пользователя
        user_id = request.created_by
        user = db.query(User).filter(User.id == user_id).first()
        user_name = user.full_name if user else "Неизвестный пользователь"
        
        if user_id not in tree_data[org]['departments'][dept]['users']:
            tree_data[org]['departments'][dept]['users'][user_id] = {
                'name': user_name,
                'user_id': str(user_id),
                'imports': {},
                'total_count': 0,
                'total_amount': 0
            }
        
        # Определяем import_id
        import_id = request.import_id
        if import_id:
            import_obj = db.query(Import).filter(Import.id == import_id).first()
            import_name = f"{user_name}" + (f" - {import_obj.comment}" if import_obj and import_obj.comment else "")
        else:
            import_id = str(user_id)  # Для одиночных заявок
            import_name = f"{user_name} - одиночные заявки"
        
        import_key = str(import_id)
        if import_key not in tree_data[org]['departments'][dept]['users'][user_id]['imports']:
            tree_data[org]['departments'][dept]['users'][user_id]['imports'][import_key] = {
                'name': import_name,
                'import_id': str(import_id) if import_id else None,
                'count': 0,
                'amount': 0
            }
        
        # Обновляем счетчики
        tree_data[org]['departments'][dept]['users'][user_id]['imports'][import_key]['count'] += 1
        tree_data[org]['departments'][dept]['users'][user_id]['imports'][import_key]['amount'] += request.amount
        
        tree_data[org]['departments'][dept]['users'][user_id]['total_count'] += 1
        tree_data[org]['departments'][dept]['users'][user_id]['total_amount'] += request.amount
        
        tree_data[org]['departments'][dept]['total_count'] += 1
        tree_data[org]['departments'][dept]['total_amount'] += request.amount
        
        tree_data[org]['total_count'] += 1
        tree_data[org]['total_amount'] += request.amount
    
    # Строим дерево
    def build_tree_node(node_id, name, node_type, count, amount, **kwargs):
        return TreeNode(
            id=node_id,
            name=name,
            type=node_type,
            count=count,
            amount=amount,
            children=[],
            **kwargs
        )
    
    # Корневой узел "Все заявки"
    total_count = len(all_requests)
    total_amount = sum(req.amount for req in all_requests) if all_requests else 0
    
    root_node = build_tree_node(
        node_id="all",
        name="Все заявки",
        node_type="root",
        count=total_count,
        amount=float(total_amount)
    )
    
    # Добавляем организации
    for org_name, org_data in tree_data.items():
        org_node = build_tree_node(
            node_id=f"org_{hash(org_name)}",
            name=org_name,
            node_type="organization",
            count=org_data['total_count'],
            amount=float(org_data['total_amount']),
            organization=org_name
        )
        
        # Добавляем подразделения
        for dept_name, dept_data in org_data['departments'].items():
            dept_node = build_tree_node(
                node_id=f"dept_{hash(org_name + dept_name)}",
                name=dept_name,
                node_type="department",
                count=dept_data['total_count'],
                amount=float(dept_data['total_amount']),
                organization=org_name,
                department=dept_name
            )
            
            # Добавляем пользователей
            for user_id, user_data in dept_data['users'].items():
                user_node = build_tree_node(
                    node_id=f"user_{user_id}",
                    name=user_data['name'],
                    node_type="user",
                    count=user_data['total_count'],
                    amount=float(user_data['total_amount']),
                    user_id=user_data['user_id'],
                    organization=org_name,
                    department=dept_name
                )
                
                # Добавляем импорты
                for import_key, import_data in user_data['imports'].items():
                    import_node = build_tree_node(
                        node_id=f"import_{import_key}",
                        name=import_data['name'],
                        node_type="import",
                        count=import_data['count'],
                        amount=float(import_data['amount']),
                        import_id=import_data['import_id'],
                        user_id=user_data['user_id'],
                        organization=org_name,
                        department=dept_name
                    )
                    user_node.children.append(import_node)
                
                dept_node.children.append(user_node)
            
            org_node.children.append(dept_node)
        
        root_node.children.append(org_node)
    
    return root_node
async def get_pending_requests(
    import_id: Optional[str] = None,
    user_id: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение заявок на согласовании в казначействе
    Можно фильтровать по import_id, user_id или category
    """
    query = db.query(Request).filter(Request.status == "pending")

    # Применяем фильтрацию по категории (та же логика, что и в get_pending_categories_stats)
    if category == 'pitanie_projivanie':
        query = query.filter(
            Request.employee_category == 'pitanie_projivanie',
            Request.source == 'employee'
        )
    elif category == 'graphs':
        query = query.filter(
            Request.treasury_import_type == 'graphs',
            Request.source == 'treasury'
        )
    elif category == 'approved_by_director':
        query = query.filter(
            Request.treasury_import_type == 'approved_by_director',
            Request.source == 'treasury'
        )
    elif category == 'non_transferable':
        query = query.filter(
            Request.treasury_import_type == 'non_transferable',
            Request.source == 'treasury'
        )
    elif category == 'filialy':
        query = query.filter(
            Request.source == 'employee'
        ).filter(
            or_(
                Request.employee_category == 'filialy',
                Request.employee_category.is_(None)
            )
        )
    # Для категории 'all' не применяем дополнительных фильтров - берем все заявки
    
    if import_id:
        query = query.filter(Request.import_id == uuid.UUID(import_id))
    elif user_id:
        # Получаем все импорты пользователя и заявки из них
        user_imports = db.query(Import.id).filter(Import.user_id == uuid.UUID(user_id)).all()
        import_ids = [imp[0] for imp in user_imports]
        query = query.filter(
            Request.import_id.in_(import_ids) if import_ids else False
        )
    
    requests = query.order_by(Request.created_at.desc()).all()
    
    # Преобразуем в Response
    return [RequestResponse.from_orm(req) for req in requests]


@router.get("/pending/categories")
async def get_pending_categories_stats(
    import_id: Optional[str] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение статистики по категориям для выбранного импорта
    Используется та же логика, что и в кабинете заместителя
    """
    # Базовый запрос для заявок в статусе 'pending'
    query = db.query(Request).filter(Request.status == 'pending')

    # Отдельный запрос для категории "Все заявки"
    # Если передан import_id, то считаем только заявки этого импорта
    # Если import_id не передан, то считаем все заявки
    all_query = db.query(Request).filter(Request.status == 'pending')

    if import_id:
        query = query.filter(Request.import_id == uuid.UUID(import_id))
        # Для категории "Все заявки" ТОЖЕ применяем фильтр по import_id, если он передан
        all_query = all_query.filter(Request.import_id == uuid.UUID(import_id))

    categories = []
    
    # Определяем цвета для категорий
    colors = {
        'pitanie_projivanie': '#EF4444',      # Красный
        'filialy': '#3B82F6',                 # Синий
        'graphs': '#F59E0B',                  # Оранжевый
        'approved_by_director': '#8B5CF6',    # Фиолетовый
        'non_transferable': '#10B981',        # Зеленый
        'all': '#6B7280'                      # Серый
    }
    
    # Названия категорий (как в кабинете заместителя)
    names = {
        'pitanie_projivanie': 'Питание, проживание, связь, аренда',
        'filialy': 'Подразделения',
        'graphs': 'Графики',
        'approved_by_director': 'Утверждено генеральным директором',
        'non_transferable': 'Непереносимые оплаты',
        'all': 'Все заявки'
    }

    # 1. Питание, проживание, аренда, связь (заявки от сотрудников)
    pitanie_query = query.filter(
        Request.employee_category == 'pitanie_projivanie',
        Request.source == 'employee'
    )
    pitanie_count = pitanie_query.count()
    pitanie_amount = pitanie_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'pitanie_projivanie',
        'name': names['pitanie_projivanie'],
        'count': pitanie_count,
        'amount': float(pitanie_amount),
        'color': colors['pitanie_projivanie']
    })

    # 2. Графики (особые заявки из казначейства)
    graphs_query = query.filter(
        Request.treasury_import_type == 'graphs',
        Request.source == 'treasury'
    )
    graphs_count = graphs_query.count()
    graphs_amount = graphs_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'graphs',
        'name': names['graphs'],
        'count': graphs_count,
        'amount': float(graphs_amount),
        'color': colors['graphs']
    })

    # 3. Утверждено генеральным директором (особые заявки из казначейства)
    approved_query = query.filter(
        Request.treasury_import_type == 'approved_by_director',
        Request.source == 'treasury'
    )
    approved_count = approved_query.count()
    approved_amount = approved_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'approved_by_director',
        'name': names['approved_by_director'],
        'count': approved_count,
        'amount': float(approved_amount),
        'color': colors['approved_by_director']
    })

    # 4. Непереносимые оплаты (особые заявки из казначейства)
    non_transferable_query = query.filter(
        Request.treasury_import_type == 'non_transferable',
        Request.source == 'treasury'
    )
    non_transferable_count = non_transferable_query.count()
    non_transferable_amount = non_transferable_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'non_transferable',
        'name': names['non_transferable'],
        'count': non_transferable_count,
        'amount': float(non_transferable_amount),
        'color': colors['non_transferable']
    })

    # 5. Подразделения (заявки от сотрудников, не относящиеся к питанию/проживанию)
    filialy_query = query.filter(
        Request.source == 'employee'
    ).filter(
        or_(
            Request.employee_category == 'filialy',
            Request.employee_category.is_(None)
        )
    )
    filialy_count = filialy_query.count()
    filialy_amount = filialy_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'filialy',
        'name': names['filialy'],
        'count': filialy_count,
        'amount': float(filialy_amount),
        'color': colors['filialy']
    })

    # 6. Все заявки (ВСЕ заявки со статусом pending, без фильтрации по import_id)
    all_count = all_query.count()
    all_amount = all_query.with_entities(func.coalesce(func.sum(Request.amount), 0)).scalar() or 0
    
    categories.append({
        'id': 'all',
        'name': names['all'],
        'count': all_count,
        'amount': float(all_amount),
        'color': colors['all']
    })

    return categories
@router.post("/pending/pivot-table")
async def get_pending_pivot_table(
    data: dict,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение сводной таблицы для заявок на согласовании
    Аналог endpoint'а заместителя, но с фильтрацией по импорту
    """
    category = data.get('category', 'filialy')
    import_id = data.get('import_id')

    # Базовый запрос для заявок в статусе 'pending'
    query = db.query(Request).filter(Request.status == 'pending')

    if import_id:
        query = query.filter(Request.import_id == uuid.UUID(import_id))

    # Применяем фильтрацию по категории (та же логика, что и в get_pending_categories_stats) (та же логика, что и в get_pending_categories_stats) (та же логика, что и в get_pending_categories_stats)
    if category == 'pitanie_projivanie':
        query = query.filter(
            Request.employee_category == 'pitanie_projivanie',
            Request.source == 'employee'
        )
    elif category == 'graphs':
        query = query.filter(
            Request.treasury_import_type == 'graphs',
            Request.source == 'treasury'
        )
    elif category == 'approved_by_director':
        query = query.filter(
            Request.treasury_import_type == 'approved_by_director',
            Request.source == 'treasury'
        )
    elif category == 'non_transferable':
        query = query.filter(
            Request.treasury_import_type == 'non_transferable',
            Request.source == 'treasury'
        )
    elif category == 'filialy':
        query = query.filter(
            Request.source == 'employee'
        ).filter(
            or_(
                Request.employee_category == 'filialy',
                Request.employee_category.is_(None)
            )
        )
    # Для категории 'all' не применяем дополнительных фильтров - берем все заявки

    requests = query.all()

    # Собираем данные для сводной таблицы
    pivot_data = {}
    departments_set = set()

    for request in requests:
        org = request.organization
        dept = request.department

        departments_set.add(dept)

        if org not in pivot_data:
            pivot_data[org] = {}

        if dept not in pivot_data[org]:
            pivot_data[org][dept] = 0

        pivot_data[org][dept] += request.amount

    # Преобразуем в формат для фронтенда
    departments = sorted(list(departments_set))
    rows = []

    for org, depts in pivot_data.items():
        row = {
            'organization': org,
            'departments': {}
        }

        for dept in departments:
            row['departments'][dept] = depts.get(dept, 0)

        rows.append(row)

    # Сортируем строки по названию организации
    rows.sort(key=lambda x: x['organization'])

    return {
        'departments': departments,
        'rows': rows
}

@router.post("/pending/send-to-deputy")
async def send_to_deputy(
    request_ids: List[uuid.UUID],
    treasury_comment: str = Form(...),
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Отправка заявок на финальное согласование заместителю
    Автоматически отклоняет оставшиеся заявки в тех же импортах
    """
    if not request_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не выбраны заявки для отправки"
        )
    
    # Проверяем, что все заявки в статусе 'pending'
    requests = db.query(Request).filter(Request.id.in_(request_ids)).all()

    if len(requests) != len(request_ids):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Некоторые заявки не найдены"
        )

    for request in requests:
        if request.status != 'pending':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Заявка {request.id} имеет статус {request.status}, а должен быть 'pending'"
            )

    # Получаем все уникальные import_id из выбранных заявок
    import_ids = set()
    for request in requests:
        if request.import_id:
            import_ids.add(request.import_id)

    # Группируем заявки по категориям для создания процессов согласования
    from collections import defaultdict
    categories_dict = defaultdict(list)

    for request in requests:
        # Определяем категорию для группировки (используем ту же логику, что и в get_pending_categories_stats)
        if request.source == 'employee':
            if request.employee_category == 'pitanie_projivanie':
                category = 'pitanie_projivanie'
            else:
                category = 'filialy'
        else:
            # Для заявок из казначейства используем treasury_import_type как категорию
            category = request.treasury_import_type or 'other'
        
        categories_dict[category].append(request.id)

    created_processes = []

    # Находим заместителя (один для всех категорий)
    deputy = db.query(User).filter(
        User.role == "deputy_director",
        User.is_active == True
    ).first()

    if not deputy:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Не найден активный заместитель"
        )
    
    # Создаем ApprovalProcess для каждой категории
    for category, cat_request_ids in categories_dict.items():
        

        # Создаем процесс согласования
        approval_process = ApprovalProcess(
            id=uuid.uuid4(),
            deputy_id=deputy.id,
            category=category,
            comment=treasury_comment,
            treasury_comment=treasury_comment,
            treasury_user_id=current_user.id,
            status='pending',
            request_ids=cat_request_ids
        )

        db.add(approval_process)
        created_processes.append(approval_process)

        # Обновляем статус заявок и привязываем к процессу
        for req_id in cat_request_ids:
            request = db.query(Request).filter(Request.id == req_id).first()
            if request:
                request.status = 'approved_for_payment'
                request.approval_process_id = approval_process.id

    
    # Создаем дополнительный общий процесс, если отправляются заявки из разных категорий
    # Всегда создаем общий процесс при отправке из TreasuryPending
    if True:
        # Создаем общий процесс с категорией 'general'
        general_approval_process = ApprovalProcess(
            id=uuid.uuid4(),
            deputy_id=deputy.id,
            category='general',
            comment=treasury_comment,
            treasury_comment=treasury_comment,
            treasury_user_id=current_user.id,
            status='pending',
            request_ids=request_ids  # Все выбранные заявки
        )
        
        db.add(general_approval_process)
        created_processes.append(general_approval_process)
        
        print(f'Создан общий процесс согласования для {len(request_ids)} заявок из {len(categories_dict)} категорий')


# АВТОМАТИЧЕСКОЕ ОТКЛОНЕНИЕ ОСТАВШИХСЯ ЗАЯВОК
    rejected_count = 0
    
    if import_ids:
        # Находим все заявки в тех же импортах, которые имеют статус 'pending'
        # и НЕ входят в список выбранных заявок
        query = db.query(Request).filter(
            Request.status == 'pending',
            Request.import_id.in_(list(import_ids)),
            ~Request.id.in_(request_ids)
        )
        
        rejected_count = query.update(
            {"status": "rejected"},
            synchronize_session=False
        )
    
    db.commit()

    # Отправляем уведомления заместителю
    from app.routes.notifications import create_batch_for_approval_notification
    for process in created_processes:
        # Рассчитываем общую сумму для этой категории
        total_amount = sum([r.amount for r in requests if r.id in process.request_ids])

        create_batch_for_approval_notification(
            db=db,
            deputy_id=process.deputy_id,
            import_id=None,
            request_count=len(process.request_ids),
            categories=[process.category],
            total_amount=total_amount,
            imported_by_user=current_user
        )

    return {
        "message": f"{len(request_ids)} заявок отправлено на финальное согласование, {rejected_count} заявок отклонено",
        "created_processes": len(created_processes),
        "rejected_count": rejected_count
    }

@router.get("/approved", response_model=List[RequestResponse])
async def get_approved_requests_alias(
    category: Category = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Алиас для /approved-for-payment
    """
    return await get_approved_requests_for_payment(
        category=category,
            comment=treasury_comment,
        start_date=start_date,
        end_date=end_date,
        current_user=current_user,
        db=db
    )

@router.get("/approval-comments/{request_id}", response_model=ApprovalCommentResponse)
async def get_approval_comment_for_request(
    request_id: UUID,
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение комментария заместителя для конкретной заявки
    """
    # Находим заявку
    request = db.query(Request).filter(Request.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    
    # Если нет процесса согласования
    if not request.approval_process_id:
        return ApprovalCommentResponse(
            has_comment=False,
            treasury_comment=None,
            approval_process_id=None,
            comment=None
        )
    
    # Находим процесс согласования
    approval_process = db.query(ApprovalProcess).filter(
        ApprovalProcess.id == request.approval_process_id
    ).first()
    
    if not approval_process:
        return ApprovalCommentResponse(
            has_comment=False,
            treasury_comment=None,
            approval_process_id=None,
            comment=None
        )
    
    return ApprovalCommentResponse(
        has_comment=True,
        treasury_comment=approval_process.treasury_comment,
        approval_process_id=approval_process.id,
        comment=approval_process.comment
    )


@router.get("/batch-approval-comments", response_model=List[ApprovalCommentResponse])
async def get_batch_approval_comments(
    request_ids: List[UUID] = Query(..., description="Список ID заявок"),
    current_user: User = Depends(require_treasury),
    db: Session = Depends(get_db)
):
    """
    Получение комментариев заместителя для нескольких заявок
    """
    if not request_ids:
        return []
    
    # Находим все заявки
    requests = db.query(Request).filter(Request.id.in_(request_ids)).all()
    
    # Собираем ID процессов согласования
    approval_process_ids = [r.approval_process_id for r in requests if r.approval_process_id]
    
    if not approval_process_ids:
        return [ApprovalCommentResponse(
            has_comment=False,
            treasury_comment=None,
            approval_process_id=None,
            comment=None
        ) for _ in requests]
    
    # Находим все процессы согласования
    approval_processes = db.query(ApprovalProcess).filter(
        ApprovalProcess.id.in_(approval_process_ids)
    ).all()
    
    # Создаем словарь для быстрого доступа
    process_dict = {str(p.id): p for p in approval_processes}
    
    # Формируем ответ
    result = []
    for request in requests:
        if request.approval_process_id and str(request.approval_process_id) in process_dict:
            process = process_dict[str(request.approval_process_id)]
            result.append(ApprovalCommentResponse(
                has_comment=True,
                treasury_comment=process.treasury_comment,
                approval_process_id=process.id,
                comment=process.comment
            ))
        else:
            result.append(ApprovalCommentResponse(
                has_comment=False,
                treasury_comment=None,
                approval_process_id=None,
                comment=None
            ))
    
    return result
